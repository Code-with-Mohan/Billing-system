import tkinter as tk
from tkinter import messagebox, ttk
from tkinter import filedialog
import datetime
import os
from openpyxl import Workbook, load_workbook
from twilio.rest import Client
from PIL import Image, ImageTk

# Dummy Twilio credentials (replace with real credentials)This account belong's to code with mohan
TWILIO_ACCOUNT_SID = 'AC3b47dd5b83e98cc259d21164b7b12ace'
TWILIO_AUTH_TOKEN = 'be8845ff2a6c674d5d5de1aca723768e'
TWILIO_PHONE_NUMBER = '+12563611149'

# Password for secure operations
PASSWORD = "mnr@2024"

# File to store day sales
DAY_SALE_FILE = "day_sales.xlsx"

class BillingSoftware:
    def __init__(self, root):
        self.root = root
        self.root.title("Five Star Billing Software")
        self.root.geometry("800x600")
        
        # Add icon
        self.root.iconphoto(False, tk.PhotoImage(file="logo.png"))
        
        # Background Image
        bg_image = Image.open("background.jpg")
        bg_photo = ImageTk.PhotoImage(bg_image)
        self.bg_label = tk.Label(root, image=bg_photo)
        self.bg_label.place(x=0, y=0, relwidth=1, relheight=1)
        self.root.bg_photo = bg_photo  # Keep a reference to avoid garbage collection

        # Initialize sales data
        self.init_sales_data()

        # Password Prompt
        self.prompt_password()

        # Handle window close event
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def init_sales_data(self):
        if not os.path.exists(DAY_SALE_FILE):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sales"
            sheet.append(["Date", "Total Sale"])
            workbook.save(DAY_SALE_FILE)

    def prompt_password(self):
        self.password_window = tk.Toplevel(self.root)
        self.password_window.geometry("300x200")
        self.password_window.title("Enter Password")
        self.password_window.config(bg="#3A5FCD")
        self.password_window.resizable(False, False)
        self.password_window.grab_set()

        logo_image = Image.open("logo.png")
        logo_image = logo_image.resize((50, 50), Image.LANCZOS)
        logo_photo = ImageTk.PhotoImage(logo_image)
        tk.Label(self.password_window, image=logo_photo, bg="#3A5FCD").pack(pady=10)
        self.password_window.logo_photo = logo_photo  # Keep a reference to avoid garbage collection

        tk.Label(self.password_window, text="Password:", bg="#3A5FCD", fg="white", font=("Arial", 12)).pack(pady=10)
        self.password_entry = tk.Entry(self.password_window, show='*', font=("Arial", 12))
        self.password_entry.pack(pady=5)

        tk.Button(self.password_window, text="Submit", command=self.check_password, bg="green", fg="white", font=("Arial", 12)).pack(pady=10)

    def check_password(self):
        if self.password_entry.get() == PASSWORD:
            self.password_window.destroy()
            self.create_widgets()
        else:
            messagebox.showerror("Error", "Invalid Password")

    def create_widgets(self):
        # Title and Date/Time
        title_frame = tk.Frame(self.root, bg="#3A5FCD")
        title_frame.pack(side=tk.TOP, fill=tk.X)

        # Logo
        logo_image = Image.open("logo.png")
        logo_image = logo_image.resize((50, 50), Image.LANCZOS)
        logo_photo = ImageTk.PhotoImage(logo_image)
        tk.Label(title_frame, image=logo_photo, bg="#3A5FCD").pack(side=tk.LEFT, padx=10)
        self.root.logo_photo = logo_photo  # Keep a reference to avoid garbage collection

        tk.Label(title_frame, text="FIVE STAR BILLING SOFTWARE", font=("Arial", 24), bg="#3A5FCD", fg="white").pack(side=tk.LEFT, padx=20)
        self.time_label = tk.Label(title_frame, text="", font=("Arial", 14), bg="#3A5FCD", fg="white")
        self.time_label.pack(side=tk.RIGHT, padx=20)
        self.update_time()

        # Customer Information
        customer_frame = tk.Frame(self.root, bg="#ADD8E6")
        customer_frame.pack(fill=tk.X)
        tk.Label(customer_frame, text="Customer name:", bg="#ADD8E6").pack(side=tk.LEFT, padx=10)
        self.customer_name = tk.Entry(customer_frame)
        self.customer_name.pack(side=tk.LEFT, padx=10)
        tk.Label(customer_frame, text="Phone no:", bg="#ADD8E6").pack(side=tk.LEFT, padx=10)
        self.phone_number = tk.Entry(customer_frame)
        self.phone_number.pack(side=tk.LEFT, padx=10)

        # Clear and Refresh Customer Info Buttons
        tk.Button(customer_frame, text="Clear", command=self.clear_customer_info, bg="red", fg="white").pack(side=tk.LEFT, padx=10)
        tk.Button(customer_frame, text="Refresh", command=self.refresh_software, bg="green", fg="white").pack(side=tk.LEFT, padx=10)

        # Main Frame
        self.main_frame = tk.Frame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        self.menu_frame = tk.Frame(self.main_frame)
        self.menu_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Home Button
        tk.Button(self.menu_frame, text="Home", command=self.go_home, bg="blue", fg="white", font=("Arial", 16), width=10, height=2).pack(fill=tk.X, padx=20, pady=5)
        
        tk.Label(self.menu_frame, text="MENU:", font=("Arial", 18)).pack(anchor=tk.W, padx=10, pady=10)

        self.menu_items = {
            "Classic Chicken": [("Mini bucket 5pcs", "images/Mini bucket.png", 385), ("Big Bucket (10pcs)", "images/BIG BUCKET.png", 765),("One Piece", "images/One piece.png", 80),("Krusty bites(8 Pcs)", "images/Krusty bites.png", 80),("Peri peri chewings(4 Pcs)", "images/peri peri chewings.png", 100),("Chicken65", "images/Chicken 65.png", 75),("Add-on (mayonnaise) & (seasoning)", "images/Mayonnaise-Recipe-11.png",5)],
            "Veggie Suprise": [("Paneer delight burger", "images/paneer delight burger.png", 89), ("Veg Burger", "images/Veg Burger.png", 65),("Pizza pockets (4 Pcs)", "images/Pizza pockets.png", 59),("Veg roll", "images/veg roll.png", 65),("Veg fingers (4 Pcs)", "images/veg roll.png", 49)],
            "Grab & Eat": [("Star burger", "images/star burger.png", 60),("Hungry bird burger", "images/hungry burger.png", 85),("Tandoori Burger", "images/tanddor burger.png", 90),("Hot Crispy Burger", "images/Hot Crispy Burger.png", 110),("Chicken roll", "images/chicken roll.png", 60),("Seekh roll", "images/seekh roll.png", 70),("Tandoori roll", "images/tandoor roll.png", 75),("Krisper roll", "images/krisper roll.png", 85)],
            "Snack Attack": [("Masala fries Regular", "images/Masala fries.png", 45), ("Masala fries Large", "images/Masala fries.png", 85),("Chicken Fingers Regular", "images/chicken fingers.png", 50),("Chicken Fingers Large", "images/chicken fingers.png", 80),("Chicken Nuggets Regular", "images/Chicken nuggets.png",50),("Chicken Nuggets Large ", "images/Chicken nuggets.png", 85),("Chicken Popcorn Regular", "images/chicken popcorn.png",70),("Chicken Popcorn Large ", "images/chicken popcorn.png",120),("Cheese/Hot shots Regular", "images/chicken cheese shots.png",75),("Cheese/Hot shots Large", "images/chicken cheese shots.png",120),("Chicken Strips Regular", "images/chicken strips.png",90),("Chicken Strips Regular", "images/chicken strips.png",90),("Chicken Strips  Large", "images/chicken strips.png",145)],
            "Koli Hut": [("Chicken kebab", "images/chicken kebab.png", 70), ("Chicken chettinaad", "images/chicken chettinad.png", 70), ("Pepper chicken", "images/pepper chicken.png", 70),("Chicken masala fries", "images/chicken masala fries.png", 50),("Krunchy munchy", "images/krucnhy munchy.png", 90),("Krunchy munchy roll", "images/krucnhy munchy roll.png", 95),("Hot chicken shots", "images/hot chicken shots.png", 75),("kalmi kebab", "images/Kalmi kebab.png", 65)],
        }

        for item in self.menu_items:
            btn = tk.Button(self.menu_frame,text=item, bg="yellow", font=("Arial", 12), command=lambda i=item: self.show_items(i))
            btn.pack(fill=tk.X, padx=20, pady=5)

        button_frame = tk.Frame(self.menu_frame)
        button_frame.pack(pady=10)

        tk.Button(button_frame, text="ADD ITEM", command=self.add_item, bg="red", fg="white").pack(side=tk.LEFT, padx=10)
        tk.Button(button_frame, text="REMOVE ITEM", command=self.remove_item, bg="red", fg="white").pack(side=tk.LEFT, padx=10)
        tk.Button(button_frame, text="DAY SALE", command=self.show_day_sale, bg="red", fg="white").pack(side=tk.LEFT, padx=10)

        self.bill_frame = tk.Frame(self.main_frame)
        self.bill_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        tk.Label(self.bill_frame, text="BILL", font=("Arial", 18)).pack(anchor=tk.W, padx=10, pady=10)

        self.bill_text = tk.Text(self.bill_frame, width=40, height=20)
        self.bill_text.pack(padx=10, pady=10)

        action_frame = tk.Frame(self.bill_frame)
        action_frame.pack(pady=10)

        tk.Button(action_frame, text="CANCEL ORDER", command=self.cancel_order, bg="red", fg="white", font=("Arial", 12)).pack(side=tk.LEFT, padx=10)
        tk.Button(action_frame, text="CONFIRM ORDER", command=self.confirm_order, bg="green", fg="white", font=("Arial", 12)).pack(side=tk.LEFT, padx=10)
        tk.Button(action_frame, text="CLEAR BILL", command=self.clear_bill, bg="blue", fg="white", font=("Arial", 12)).pack(side=tk.LEFT, padx=10)
        
        # Bill Total Button
        tk.Button(action_frame, text="BILL TOTAL", command=self.calculate_total, bg="purple", fg="white", font=("Arial", 12)).pack(side=tk.LEFT, padx=10)

        # Change Calculation
        tk.Label(self.bill_frame, text="Amount Received:", font=("Arial", 12)).pack(anchor=tk.W, padx=10, pady=5)
        self.amount_received_entry = tk.Entry(self.bill_frame, font=("Arial", 12))
        self.amount_received_entry.pack(anchor=tk.W, padx=10, pady=5)
        tk.Button(self.bill_frame, text="Calculate Change", command=self.calculate_change, bg="orange", fg="white", font=("Arial", 12)).pack(anchor=tk.W, padx=10, pady=5)

        # Footer
        footer_frame = tk.Frame(self.root, bg="#3A5FCD")
        footer_frame.pack(side=tk.BOTTOM, fill=tk.X)
        tk.Label(footer_frame, text="© MNR", font=("Arial", 12), bg="#3A5FCD", fg="white").pack(side=tk.RIGHT, padx=20)

        self.item_frame = None

    def update_time(self):
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.time_label.config(text=now)
        self.root.after(1000, self.update_time)

    def show_items(self, category):
        if self.item_frame:
            self.item_frame.destroy()

        self.item_frame = tk.Frame(self.main_frame)
        self.item_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        items = self.menu_items[category]

        for item in items:
            item_name, item_image, item_price = item
            item_frame = tk.Frame(self.item_frame)
            item_frame.pack(fill=tk.X, pady=5)

            img = Image.open(item_image)
            img = img.resize((50, 50), Image.LANCZOS)
            img = ImageTk.PhotoImage(img)

            tk.Label(item_frame, image=img).pack(side=tk.LEFT, padx=10)
            tk.Label(item_frame, text=f"{item_name}\n₹{item_price:.2f}", font=("Arial", 12)).pack(side=tk.LEFT, padx=10)

            select_button = tk.Button(item_frame, text="Select Item", command=lambda n=item_name, p=item_price: self.add_to_bill(n, p))
            select_button.pack(side=tk.RIGHT, padx=10)

            remove_button = tk.Button(item_frame, text="Remove Item", command=lambda n=item_name, c=category: self.remove_from_menu(n, c), bg="red", fg="white")
            remove_button.pack(side=tk.RIGHT, padx=10, pady=5)

            item_frame.img = img  # Keep a reference to the image to avoid garbage collection

    def add_to_bill(self, item_name, item_price):
        self.bill_text.insert(tk.END, f"{item_name} - ₹{item_price:.2f}\n")

    def add_item(self):
        add_window = tk.Toplevel(self.root)
        add_window.geometry("400x400")
        add_window.title("Add Item")
        add_window.config(bg="#3A5FCD")
        add_window.resizable(False, False)
        add_window.grab_set()

        logo_image = Image.open("logo.png")
        logo_image = logo_image.resize((50, 50), Image.LANCZOS)
        logo_photo = ImageTk.PhotoImage(logo_image)
        tk.Label(add_window, image=logo_photo, bg="#3A5FCD").pack(pady=10)
        add_window.logo_photo = logo_photo  # Keep a reference to avoid garbage collection

        tk.Label(add_window, text="Category:", bg="#3A5FCD", fg="white").pack(pady=5)
        category_entry = tk.Entry(add_window)
        category_entry.pack(pady=5)

        tk.Label(add_window, text="Item Name:", bg="#3A5FCD", fg="white").pack(pady=5)
        item_name_entry = tk.Entry(add_window)
        item_name_entry.pack(pady=5)

        tk.Label(add_window, text="Item Price:", bg="#3A5FCD", fg="white").pack(pady=5)
        item_price_entry = tk.Entry(add_window)
        item_price_entry.pack(pady=5)

        tk.Label(add_window, text="Item Image:", bg="#3A5FCD", fg="white").pack(pady=5)
        item_image_path = tk.StringVar()
        tk.Entry(add_window, textvariable=item_image_path).pack(pady=5)
        tk.Button(add_window, text="Browse", command=lambda: self.browse_image(item_image_path)).pack(pady=5)

        def add_new_item():
            category = category_entry.get()
            item_name = item_name_entry.get()
            item_price = item_price_entry.get()
            item_image = item_image_path.get()

            if category and item_name and item_price and item_image:
                if category not in self.menu_items:
                    self.menu_items[category] = []
                self.menu_items[category].append((item_name, item_image, float(item_price)))
                add_window.destroy()
                messagebox.showinfo("Success", "Item added successfully")
                self.refresh_menu()
            else:
                messagebox.showerror("Error", "All fields are required")

        tk.Button(add_window, text="Add", command=add_new_item, bg="green", fg="white").pack(pady=10)

    def browse_image(self, image_path):
        file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
        if file_path:
            image_path.set(file_path)

    def remove_item(self):
        remove_window = tk.Toplevel(self.root)
        remove_window.geometry("300x200")
        remove_window.title("Remove Item")
        remove_window.config(bg="#3A5FCD")
        remove_window.resizable(False, False)
        remove_window.grab_set()

        logo_image = Image.open("logo.png")
        logo_image = logo_image.resize((50, 50), Image.LANCZOS)
        logo_photo = ImageTk.PhotoImage(logo_image)
        tk.Label(remove_window, image=logo_photo, bg="#3A5FCD").pack(pady=10)
        remove_window.logo_photo = logo_photo  # Keep a reference to avoid garbage collection

        tk.Label(remove_window, text="Category:", bg="#3A5FCD", fg="white").pack(pady=5)
        category_entry = tk.Entry(remove_window)
        category_entry.pack(pady=5)

        tk.Label(remove_window, text="Item Name:", bg="#3A5FCD", fg="white").pack(pady=5)
        item_name_entry = tk.Entry(remove_window)
        item_name_entry.pack(pady=5)

        def remove_existing_item():
            category = category_entry.get()
            item_name = item_name_entry.get()

            if category in self.menu_items and item_name:
                self.menu_items[category] = [item for item in self.menu_items[category] if item[0] != item_name]
                remove_window.destroy()
                messagebox.showinfo("Success", "Item removed successfully")
                self.refresh_menu()
            else:
                messagebox.showerror("Error", "Invalid category or item name")

        tk.Button(remove_window, text="Remove", command=remove_existing_item, bg="red", fg="white").pack(pady=10)

    def remove_from_menu(self, item_name, category):
        if category in self.menu_items and item_name:
            self.menu_items[category] = [item for item in self.menu_items[category] if item[0] != item_name]
            messagebox.showinfo("Success", "Item removed successfully")
            self.refresh_menu()

    def refresh_menu(self):
        # Clear existing menu items
        for widget in self.menu_frame.winfo_children():
            widget.destroy()

        # Add home button
        tk.Button(self.menu_frame, text="Home", command=self.go_home, bg="blue", fg="white", font=("Arial", 16), width=10, height=2).pack(fill=tk.X, padx=20, pady=5)
        
        tk.Label(self.menu_frame, text="MENU:", font=("Arial", 18)).pack(anchor=tk.W, padx=10, pady=10)

        # Recreate menu items
        for item in self.menu_items:
            btn = tk.Button(self.menu_frame, text=item, bg="yellow", font=("Arial", 12), command=lambda i=item: self.show_items(i))
            btn.pack(fill=tk.X, padx=20, pady=5)

        # Recreate other buttons
        button_frame = tk.Frame(self.menu_frame)
        button_frame.pack(pady=10)

        tk.Button(button_frame, text="ADD ITEM", command=self.add_item, bg="red", fg="white").pack(side=tk.LEFT, padx=10)
        tk.Button(button_frame, text="REMOVE ITEM", command=self.remove_item, bg="red", fg="white").pack(side=tk.LEFT, padx=10)
        tk.Button(button_frame, text="DAY SALE", command=self.show_day_sale, bg="red", fg="white").pack(side=tk.LEFT, padx=10)

    def clear_customer_info(self):
        self.customer_name.delete(0, tk.END)
        self.phone_number.delete(0, tk.END)

    def refresh_software(self):
        self.root.destroy()
        root = tk.Tk()
        app = BillingSoftware(root)
        root.mainloop()

    def go_home(self):
        if self.item_frame:
            self.item_frame.destroy()
        self.main_frame.pack_forget()
        self.main_frame.pack(fill=tk.BOTH, expand=True)

    def show_day_sale(self):
        day_sale_window = tk.Toplevel(self.root)
        day_sale_window.geometry("300x200")
        day_sale_window.title("Day Sale")
        day_sale_window.config(bg="#3A5FCD")
        day_sale_window.resizable(False, False)
        day_sale_window.grab_set()

        logo_image = Image.open("logo.png")
        logo_image = logo_image.resize((50, 50), Image.LANCZOS)
        logo_photo = ImageTk.PhotoImage(logo_image)
        tk.Label(day_sale_window, image=logo_photo, bg="#3A5FCD").pack(pady=10)
        day_sale_window.logo_photo = logo_photo  # Keep a reference to avoid garbage collection

        workbook = load_workbook(DAY_SALE_FILE)
        sheet = workbook.active
        total_sales = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            total_sales += row[1]
        workbook.close()

        tk.Label(day_sale_window, text=f"Total Sales for the Day: ₹{total_sales:.2f}", bg="#3A5FCD", fg="white", font=("Arial", 12)).pack(pady=10)

    def cancel_order(self):
        cancel_window = tk.Toplevel(self.root)
        cancel_window.geometry("300x200")
        cancel_window.title("Enter Password to Cancel Order")
        cancel_window.config(bg="#3A5FCD")
        cancel_window.resizable(False, False)
        cancel_window.grab_set()

        logo_image = Image.open("logo.png")
        logo_image = logo_image.resize((50, 50), Image.LANCZOS)
        logo_photo = ImageTk.PhotoImage(logo_image)
        tk.Label(cancel_window, image=logo_photo, bg="#3A5FCD").pack(pady=10)
        cancel_window.logo_photo = logo_photo  # Keep a reference to avoid garbage collection

        tk.Label(cancel_window, text="Password:", bg="#3A5FCD", fg="white", font=("Arial", 12)).pack(pady=10)
        cancel_password_entry = tk.Entry(cancel_window, show='*', font=("Arial", 12))
        cancel_password_entry.pack(pady=5)

        def confirm_cancel():
            if cancel_password_entry.get() == PASSWORD:
                cancel_window.destroy()
                self.bill_text.delete(1.0, tk.END)
                messagebox.showinfo("Success", "Order Cancelled")
            else:
                messagebox.showerror("Error", "Invalid Password")

        tk.Button(cancel_window, text="Submit", command=confirm_cancel, bg="green", fg="white", font=("Arial", 12)).pack(pady=10)

    def confirm_order(self):
        customer_phone = self.phone_number.get()
        if customer_phone:
            bill_details = self.bill_text.get(1.0, tk.END).strip()
            if bill_details:
                self.send_sms(customer_phone, bill_details)
                self.save_sale(bill_details)
                messagebox.showinfo("Success", "Order Confirmed and Bill Sent")
                self.bill_text.delete(1.0, tk.END)
            else:
                messagebox.showerror("Error", "Bill is empty")
        else:
            messagebox.showerror("Error", "Phone number is required")

    def send_sms(self, to_phone, bill_details):
        client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
        client.messages.create(
            to=to_phone,
            from_=TWILIO_PHONE_NUMBER,
            body=f"Your Bill:\n{bill_details}"
        )

    def save_sale(self, bill_details):
        today = datetime.datetime.now().strftime("%Y-%m-%d")
        total_sale = sum(float(line.split('- ₹')[-1]) for line in bill_details.strip().split('\n'))
        workbook = load_workbook(DAY_SALE_FILE)
        sheet = workbook.active
        sheet.append([today, total_sale])
        workbook.save(DAY_SALE_FILE)

    def clear_bill(self):
        self.bill_text.delete(1.0, tk.END)

    def calculate_total(self):
        bill_details = self.bill_text.get(1.0, tk.END).strip()
        total = sum(float(line.split('- ₹')[-1]) for line in bill_details.strip().split('\n'))
        messagebox.showinfo("Total Bill", f"Total Bill: ₹{total:.2f}")

    def calculate_change(self):
        try:
            total_bill = sum(float(line.split('- ₹')[-1]) for line in self.bill_text.get(1.0, tk.END).strip().split('\n'))
            amount_received = float(self.amount_received_entry.get())
            change = amount_received - total_bill
            messagebox.showinfo("Change", f"Change to be returned: ₹{change:.2f}")
        except ValueError:
            messagebox.showerror("Error", "Invalid input")

    def on_closing(self):
        if messagebox.askokcancel("Quit", "Do you want to save changes before exiting?"):
            self.save_sale(self.bill_text.get(1.0, tk.END).strip())
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = BillingSoftware(root)
    root.mainloop()
